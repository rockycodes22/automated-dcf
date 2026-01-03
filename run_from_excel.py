#imports
import pandas as pd
from dcf_model import run_dcf
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule

#pandas dataframe
df = pd.read_excel("dcf_inputs.xlsx", sheet_name="inputs", header=None)
df.columns = ["label", "value"]



ticker = df.loc[df["label"] == "Ticker", "value"].values[0]


#------------------------
#ASSUMPTIONS FROM EXCEL
#------------------------
assumptions = {
    "wacc": df.loc[df["label"] == "WACC", "value"].values[0],
    "terminal_g": df.loc[df["label"] == "Terminal Growth", "value"].values[0],
    "tax_rate": df.loc[df["label"] == "Tax Rate", "value"].values[0],
    "forecast_years": int(df.loc[df["label"] == "Forecast Years", "value"].values[0]),
    "da_percentage": df.loc[df["label"] == "D&A %", "value"].values[0],
    "capex_percentage": df.loc[df["label"] == "CapEx %", "value"].values[0],
    "nwc_percentage": df.loc[df["label"] == "NWC %", "value"].values[0],
    "cash": df.loc[df["label"] == "Cash", "value"].values[0],
    "debt": df.loc[df["label"] == "Debt", "value"].values[0],
    "shares_outstanding": df.loc[df["label"] == "Shares Outstanding", "value"].values[0], 
    "wacc_range": [float(x) for x in df.loc[df["label"] == "WACC Range", "value"].values[0].split(",")],
    "terminal_g_range": [float(x) for x in df.loc[df["label"] == "Terminal g Range", "value"].values[0].split(",")],
}

results = run_dcf(ticker, assumptions)
print(results["summary"])

#------------------------
#EXPORT TO EXCEL
#------------------------

summary_df = pd.DataFrame(
    list(results["summary"].items()),
    columns=["Metric", "Value"]
)

with pd.ExcelWriter(
    "dcf_inputs.xlsx",
    engine="openpyxl",
    mode="a",
    if_sheet_exists="replace"
) as writer:
    summary_df.to_excel(
        writer,
        sheet_name="output",
        index=False
    )

# ------------------------
# EXPORT SENSITIVITY TABLE
# ------------------------

sens_df = results["valuation"]["sensitivity"]

with pd.ExcelWriter(
    "dcf_inputs.xlsx",
    engine="openpyxl",
    mode="a",
    if_sheet_exists="replace"
) as writer:
    sens_df.to_excel(
        writer,
        sheet_name="sensitivity"
    )

# ------------------------
# FORMAT EXCEL OUTPUT
# ------------------------

wb = load_workbook("dcf_inputs.xlsx")
ws = wb["output"]

# Format Value column as currency
for row in range(2, ws.max_row + 1):
    cell = ws[f"B{row}"]

    if isinstance(cell.value, (int, float)):
        if cell.value > 1_000_000:
            cell.number_format = '$#,##0'
        else:
            cell.number_format = '$#,##0.00'

# Auto-size columns 
for column in ws.columns:
    max_length = 0
    col_letter = column[0].column_letter

    for cell in column:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))

    ws.column_dimensions[col_letter].width = max_length + 2

wb.save("dcf_inputs.xlsx")

# ------------------------
# FORMAT SENSITIVITY TABLE
# ------------------------
ws = wb["sensitivity"]

# Auto-size columns
for column in ws.columns:
    max_length = 0
    col_letter = column[0].column_letter
    for cell in column:
        if cell.value is not None:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

# Currency format for body 
for row in ws.iter_rows(min_row=2, min_col=2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0'

# Optional: color scale (green = higher EV, red = lower EV)
last_row = ws.max_row
last_col = ws.max_column
color_rule = ColorScaleRule(
    start_type='min', start_color='F8696B',
    mid_type='percentile', mid_value=50, mid_color='FFEB84',
    end_type='max', end_color='63BE7B'
)

ws.conditional_formatting.add(
    f"B2:{ws.cell(row=last_row, column=last_col).coordinate}",
    color_rule
)

wb.save("dcf_inputs.xlsx")
